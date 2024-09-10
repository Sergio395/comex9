from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Cargar el archivo Excel
file_path = 'G:/Mi unidad/001ComercioExterior/Planilla_cobro_ME.xlsx'
workbook = load_workbook(file_path)

# Seleccionar la hoja de trabajo
sheet = workbook.active

# Insertar una fila antes de los encabezados
sheet.insert_rows(1)

# Establecer el formato de los encabezados
header_fill = PatternFill(start_color='d9e1f2', end_color='d9e1f2', fill_type="solid")
header_font = Font(bold=True)
for cell in sheet[2]:
    cell.fill = header_fill
    cell.font = header_font

# Establecer el ancho de las columnas
for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column].width = adjusted_width

# Guardar los cambios en el archivo
workbook.save('Planilla_cobro_ME_.xlsx')
import shutil
import datetime
from pathlib import Path
import os
ruta_origen ='G:/Mi unidad/001ComercioExterior/'
ruta_destino = 'G:/Unidades compartidas/Tesoreria/COMERCIO EXTERIOR/'

shutil.copy(ruta_origen +'Planilla_cobro_ME_.xlsx' , ruta_destino)


fecha = datetime.datetime.now()
fe=f"'{fecha.day}-{fecha.month}-{fecha.year}_{fecha.hour} hs_{fecha.minute}'"

ruta2 = Path(ruta_destino+'Planilla_cobro_ME_.xlsx')
ruta2.rename(ruta_destino+'Planilla_cobro_ME_'+str(fe)+'.xlsx')
#os.rename(ruta_destino+'Planilla_cobro_ME_.xlsx','Planilla_cobro_ME_'+str(fe)+'.xlsx')

#print(fe)
print("Archivo modificado guardado correctamente.")
